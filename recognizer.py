import cv2
from face_detection import face
from keras.models import load_model
import numpy as np
from embedding import emb
from datetime import datetime, date
import openpyxl,os
import xlrd

name={0:"Aishwarya",1:"Akshata",2:"Ameya",3:"Anjali",4:"Ankit",5:"Bhagyashri",6:"Chinmay",7:"Dipak",8:"Harsh",9:"KajalH",10:"KajalK",11:"Kanchan",12:"Kumudini",13:"Monika",14:"Neha",15:"Omkar",16:"Pravin",17:"Priya",18:"Sachin",19:"Sakshi",20:"Sayali",21:"Snehanjali"}

e=emb()
fd=face()

item='P'
today = date.today()
d1 = today.strftime("%d-%m-%Y")
#print(d1)

book = openpyxl.load_workbook('attendance.xlsx')
sub=input('enter subject')
sheet = book.get_sheet_by_name(sub)  

loc = ("attendance.xlsx") 
  
wb = xlrd.open_workbook(loc) 
sh = wb.sheet_by_index(0) 
  
# Extracting number of columns 
c=sh.ncols
 
#print(sh.cell_value(0, c-1)) 
d2=sh.cell_value(0, c-1)

if d1!=d2:
    sheet.cell(row=1, column=c+1).value = d1 
    col=c+1
else:
    col=c

model=load_model('face_reco2.MODEL')

cap=cv2.VideoCapture(0)

ret,frame=cap.read()
det,coor=fd.detectFace(frame)

if(det is not None):
    for i in range(len(det)):
        detected=det[i]
        k=coor[i]
        #f=detected
        detected=cv2.resize(detected,(160,160))
        detected=detected.astype('float')/255.0
        detected=np.expand_dims(detected,axis=0)
        feed=e.calculate(detected)
        feed=np.expand_dims(feed,axis=0)
        feed=np.reshape(feed,(128,1))
        feed = np.reshape(feed, (feed.shape[1], feed.shape[0],1))
        prediction=model.predict(feed)
        print(prediction)
        result=int(np.argmax(prediction))
        print(result)
        if(np.max(prediction)>.80):
            for i in name:
                if(result==i):
                    label=name[i]
                    r=i
                    sheet.cell(row=r+2, column=col).value = item 
                    book.save('attendance.xlsx')
        else:
            label='unknown'
        

        cv2.putText(frame,label,(k[0],k[1]),cv2.FONT_HERSHEY_SIMPLEX,1,(255,255,255),2)
        
        cv2.rectangle(frame,(k[0],k[1]),(k[0]+k[2],k[1]+k[3]),(252,160,39),3)
cv2.imshow('frame',frame)
cv2.waitKey(0)    
cap.release()
cv2.destroyAllWindows()
